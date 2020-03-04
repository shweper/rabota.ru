import time

from selenium.webdriver import ActionChains

from selenium import webdriver
from openpyxl import load_workbook
import random
import xlrd
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
browser = webdriver.Chrome(chrome_options=options)
driver = browser

browser.get('https://nn.rabota.ru/v3_myVacancy.html?action=create&company_registered=true&employer_registered=true')
error_string = 1119
# chrome_options = Options()
# chrome_options.add_argument("--no-startup-window")
driver.refresh()
driver.implicitly_wait(1)


wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)
sheet = wb[lst[2]]
sheet.title
login = (sheet.cell(row=1, column=2).value)
password = (sheet.cell(row=2, column=2).value)
# Логинимся
login_bar_xpath = '//*[@id="mail"]'
login_bar = browser.find_element_by_xpath(login_bar_xpath)
login_bar.send_keys(login)

pass_bar_xpath = '//*[@id="password"]'
pass_bar = browser.find_element_by_xpath(pass_bar_xpath)
pass_bar.send_keys(password)
# кликкаем войти
driver.find_element_by_xpath('//*[@id="authForm"]/input[5]').click()

#переходим в отклики
browser.find_element_by_xpath('/html/body/div[2]/div[1]/div[2]/ul/li[2]/a').click()

#Выводим на странице 100 вакансий
def vivod_100():

    browser.find_element_by_xpath('//*[@id="global"]/div[3]/div/div[2]/div[2]/span/div[1]/a').click()
    browser.find_element_by_xpath('//*[@id="viewSettings"]/div/div/div/div[1]/div[1]').click()
    browser.find_element_by_xpath('//*[@id="viewSettings"]/div/div/div/div[2]/a[3]').click()
    browser.find_element_by_xpath('//*[@id="viewSettings"]/button').click()
vivod_100()
#while otklik == False:

#Иногда сайт просит оценить работу, может вылезти рандомно, по этому ввёл эту функцию, закрывающую баннер
def closers():
    try:
        browser.find_element_by_xpath('/html/body/div[3]/button').click()
    except:
        print("Идём дальше")
#Прокликиваем все инвайты
def invite():
    link = 0
    invite = browser.find_elements_by_class_name('link_invite')
    for link in range(len(invite)):
        print(invite[link])
        link+=1
        print(len(invite))
        print("клик клик", link)



#Основная программа кликера
def main():
    n = 0
    ns = 0
    for n in range(100):

        tr_white = browser.find_elements_by_class_name("tr_white")
        tds = tr_white[n].find_elements_by_tag_name('td')
        try:
            tds[1].find_element_by_tag_name('a')
            tds[1].find_element_by_tag_name('a').click()
            try:
                invite()
                print("TRUUU")
                browser.get('https://nn.rabota.ru/v3_vacancyResponseTable.html')
                vivod_100()
                time.sleep(1)
                try:
                    tds[2].find_element_by_tag_name('a')
                    tds[2].find_element_by_tag_name('a').click()
                    try:
                        invite()
                        print("TRUUU__2")
                        n += 1  # НА ПРОДЕ УБРАТЬ!????????????????????????????????????
                        browser.get('https://nn.rabota.ru/v3_vacancyResponseTable.html')
                        vivod_100()
                        time.sleep(1)
                    except:
                        closers()
                        n += 1
                        print('NONE____2!!')
                except:
                    closers()
                    n +=1
                    print('1')

            except:
                 closers()
                 print('NONE')
                 n += 1

        except:

            try:
                tds[2].find_element_by_tag_name('a')
                tds[2].find_element_by_tag_name('a').click()
                try:
                    invite()
                    print("TRUUU__2")
                    n += 1  # НА ПРОДЕ УБРАТЬ!????????????????????????????????????????????????
                    browser.get('https://nn.rabota.ru/v3_vacancyResponseTable.html')
                    vivod_100()
                    time.sleep(1)
                except:
                     closers()
                     n += 1
                     print('NONE____2!!')
            except:
                closers()
                print('OVER NONE')
                n +=1



main()
try:
    browser.find_element_by_xpath('//*[@id="global"]/div[4]/div/div[2]/div/a[4]').click()
except:
    browser.find_element_by_xpath('//*[@id="global"]/div[4]/div/div[2]/div/a[5]').click()

main()
try:
    browser.find_element_by_xpath('//*[@id="global"]/div[4]/div/div[2]/div/a[4]').click()
except:
    browser.find_element_by_xpath('//*[@id="global"]/div[4]/div/div[2]/div/a[5]').click()

main()
    #tr_white[n].find_element_by_tag_name('span').click()



'''
vacancy_link = browser.find_elements_by_class_name("vacancy_link ")
tr_white = browser.find_elements_by_class_name("tr_white")
span_num = tr_white[2].find_element_by_tag_name('span')
span_num.click()
print(span_num)


n = 7
while n < len(vacancy_link):
    vacancy_link = browser.find_elements_by_class_name("vacancy_link ")
    vacancy_link[n].click()
    #browser.find_element_by_xpath(vacancy_link[0]).click()
    time.sleep(1)
    try:
        browser.find_element_by_xpath('//*[@id="invite130600012"]')
        print("TRUUU")
    except:
        print('NONE')
    n+=1
    browser.get('https://nn.rabota.ru/v3_vacancyResponseTable.html')
    time.sleep(1)
'''
#xpath Вакансии с откликом
#//*[@id="global"]/div[4]/div/div[1]/div[2]/div/table/tbody/tr[3]/td[1]/a
#/html/body/div[2]/div[2]/div[2]/div/div[4]/div/div[1]/div[2]/div/table/tbody/tr[3]/td[1]/a