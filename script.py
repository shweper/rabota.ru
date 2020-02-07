from openpyxl import load_workbook
import random
from splinter import Browser
# import pandas as pd
import xlrd
wb = xlrd.open_workbook('DATA.xlsx')
sheet_xlr = wb.sheet_by_name('Вакансии')
a = 2
rows_sheet = sheet_xlr.nrows
print(rows_sheet)


wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)

# print(random.choice(lst))
# Берём рандомный лист
# sheet = wb[random.choice(lst)]
sheet = wb[lst[0]]
sheet.title

# random.randint(A, B) - случайное целое число N, A ≤ N ≤ B.
#a = 0
# b = [a]

#for row in sheet.iter_cols(max_col=1):
#    for value in row:
#        a = a+1
#print(a)

# альтернативный рандом
#        b.append(a)
# b.pop()
# b.pop()
# b.pop()
# print(b)
# i = random.choice(b)

i = random.randint(3, rows_sheet)
################### заводим содержимое в переменные ################################
# print(a)
vacancy = (sheet.cell(row=i, column=1).value)
zp_ot = (sheet.cell(row=i, column=2).value)
zp_do = (sheet.cell(row=i, column=3).value)
opyt = (sheet.cell(row=i, column=4).value)
obraz = (sheet.cell(row=i, column=5).value)
pol = (sheet.cell(row=i, column=6).value)
vozr_ot = (sheet.cell(row=i, column=7).value)
vozr_do = (sheet.cell(row=i, column=8).value)
opisanie = (sheet.cell(row=i, column=9).value)
name_hr = (sheet.cell(row=i, column=10).value)
mail = (sheet.cell(row=i, column=11).value)
kod = (sheet.cell(row=i, column=12).value)
phone = (sheet.cell(row=i, column=13).value)
priem_zv_c = (sheet.cell(row=i, column=14).value)
priem_zv_do = (sheet.cell(row=i, column=15).value)
dni_priem = (sheet.cell(row=i, column=16).value)
company = (sheet.cell(row=i, column=17).value)
opis_company = (sheet.cell(row=i, column=18).value)
# gorod = (sheet.cell(row=i, column=18).value)
# print(priem_zv_c)

###############     Browser     ###############

browser = Browser('chrome')
browser.driver.set_window_size(600, 800)
browser.visit('https://nn.rabota.ru/v3_myVacancy.html?action=create&company_registered=true&employer_registered=true')
# Логинимся
login_bar_xpath = '//*[@id="mail"]'
login_bar = browser.find_by_xpath(login_bar_xpath)[0]
login_bar.fill('shweper@ya.ru')

pass_bar_xpath = '//*[@id="password"]'
pass_bar = browser.find_by_xpath(pass_bar_xpath)[0]
pass_bar.fill('Fqtawe98')
# кликкаем войти
browser.find_by_xpath('//*[@id="authForm"]/input[5]').click()


vak_bar_xpath = '//*[@id="custom_position"]'
vak_bar = browser.find_by_xpath(vak_bar_xpath)[0]
vak_bar.fill(vacancy)

zp_ot_bar_xpath = '//*[@id="salary_from"]'
zp_ot_bar = browser.find_by_xpath(zp_ot_bar_xpath)[0]
zp_ot_bar.fill(zp_ot)

zp_do_bar_xpath = '//*[@id="salary_to"]'
zp_do_bar = browser.find_by_xpath(zp_do_bar_xpath)[0]
zp_do_bar.fill(zp_do)

###############     Опыт работы     ###############

browser.find_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[1]/td[2]').click()
if opyt == 0:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[4]').click()
elif opyt == 1:
    browser.find_by_xpath('//*[ @ id = "offer_experience_year_count"] / option[5]').click()
elif opyt == 2:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[6]').click()
elif opyt == 3:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[7]').click()
elif opyt == 4:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[8]').click()
elif opyt == 5:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[9]').click()
elif opyt == 6:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[10]').click()
elif opyt == 7:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[11]').click()
elif opyt == 8:
     browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
elif opyt == 9:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
elif opyt == 10:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
elif opyt == 100:
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()

###############     Образование     ###############

browser.find_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select').click()
if obraz == "высшее":
    browser.find_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[2]').click()
elif obraz == "неполное высшее":
    browser.find_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[3]').click()
elif obraz == "среднее профессиональное":
    browser.find_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[4]').click()
elif obraz == "среднее":
    browser.find_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[5]').click()
elif obraz == "любое":
    browser.find_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[1]').click()

###############     Пол     ###############

browser.find_by_xpath('//*[@id="is_male"]').click()
if pol == 'Не важно':
    browser.find_by_xpath('//*[@id="is_male"]/option[1]').click()
elif pol == 'Мужской':
    browser.find_by_xpath('//*[@id="is_male"]/option[2]').click()
elif pol == 'Женский':
    browser.find_by_xpath('//*[@id="is_male"]/option[3]').click()

###############     возраст     ###############

vozr_ot_bar_xpath = '//*[@id="age_from"]'
vozr_ot_bar = browser.find_by_xpath(vozr_ot_bar_xpath)[0]
vozr_ot_bar.fill(vozr_ot)

vozr_do_bar_xpath = '//*[@id="age_to"]'
vozr_do_bar = browser.find_by_xpath(vozr_do_bar_xpath)[0]
vozr_do_bar.fill(vozr_do)

###############     Описание вакинсии     ###############
#browser.find_by_xpath('//*[@id="mce_3"]/button').click()

# browser.find_by_tag('html').click()
q = browser.find_by_id('tinymce').first
q.fill('qwereewrqw')
# opisanie_bar_xpath = '//*[@id="tinymce"]'
# opisanie_bar = browser.find_by_xpath(opisanie_bar_xpath)[0]
# opisanie_bar.fill(opisanie)

###############    Контактная информация    ###############

name_hr_bar_xpath = '//*[@id="vacancyForm"]/div[5]/div/div/table/tbody/tr[1]/td[2]/div/input'
name_hr_bar = browser.find_by_xpath(name_hr_bar_xpath)[0]
name_hr_bar.fill(name_hr)


mail_bar_xpath = '//*[@id="contact_email"]'
mail_bar = browser.find_by_xpath(mail_bar_xpath)[0]
mail_bar.fill(mail)

###############    Phone    ###############

kod_bar_xpath = '//*[@id="phoneContainerCode_"]'
kod_bar = browser.find_by_xpath(kod_bar_xpath)[0]
kod_bar.fill(kod)

phone_bar_xpath = '//*[@id="phoneContainerNumber_"]'
phone_bar = browser.find_by_xpath(phone_bar_xpath)[0]
phone_bar.fill(phone)

###############    время для звонков    ###############

if priem_zv_c == '0:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[1]').mouse_over()
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[1]').click()
elif priem_zv_c == '0:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[1]').mouse_over()
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[2]').click()
elif priem_zv_c == '1:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[3]').click()
elif priem_zv_c == '1:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[4]').click()
elif priem_zv_c == '2:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[5]').click()
elif priem_zv_c == '2:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[6]').click()
elif priem_zv_c == '3:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[7]').click()
elif priem_zv_c == '3:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[8]').click()
elif priem_zv_c == '4:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[9]').click()
elif priem_zv_c == '4:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[10]').click()
elif priem_zv_c == '5:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[11]').click()
elif priem_zv_c == '5:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[12]').click()
elif priem_zv_c == '6:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[13]').click()
elif priem_zv_c == '6:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[14]').click()
elif priem_zv_c == '7:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[15]').click()
elif priem_zv_c == '7:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[16]').click()
elif priem_zv_c == '8:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[17]').click()
elif priem_zv_c == '8:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[18]').click()
elif priem_zv_c == '9:00:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[19]').click()
elif priem_zv_c == '9:30:00':
    browser.find_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[20]').click()
elif priem_zv_c == '10:00:00':
     browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
elif priem_zv_c == '10:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
elif priem_zv_c == '11:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
elif priem_zv_c == '11:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()
elif priem_zv_c == '12:00:00':
    browser.find_by_xpath('// *[ @ id = "offer_experience_year_count"] / option[5]').click()
elif priem_zv_c == '12:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[6]').click()
elif priem_zv_c == '13:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[7]').click()
elif priem_zv_c == '13:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[8]').click()
elif priem_zv_c == '14:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[9]').click()
elif priem_zv_c == '14:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[10]').click()
elif priem_zv_c == '15:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[11]').click()
elif priem_zv_c == '16:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
elif priem_zv_c == '16:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
elif priem_zv_c == '17:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
elif priem_zv_c == '17:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()
elif priem_zv_c == '18:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[4]').click()
elif priem_zv_c == '18:30:00':
    browser.find_by_xpath('// *[ @ id = "offer_experience_year_count"] / option[5]').click()
elif priem_zv_c == '19:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[6]').click()
elif priem_zv_c == '19:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[7]').click()
elif priem_zv_c == '20:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[8]').click()
elif priem_zv_c == '20:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[9]').click()
elif priem_zv_c == '21:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[10]').click()
elif priem_zv_c == '21:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[11]').click()
elif priem_zv_c == '22:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
elif priem_zv_c == '22:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
elif priem_zv_c == '23:00:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
elif priem_zv_c == '23:30:00':
    browser.find_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()

###############    дни для звонков    ###############
browser.find_by_xpath('//*[@id="phoneContainerCallPeriod_"]').click()
if opyt == 'Рабочие дни':
    browser.find_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[1]').click()
elif opyt == 'Выходные дни':
    browser.find_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[2]').click()
elif opyt == 'Любой день':
    browser.find_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[3]').click()



###############    Phone    ###############



#browser.quit()