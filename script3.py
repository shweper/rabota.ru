from openpyxl import load_workbook
import random

wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)

sheet = wb[lst[1]]
sheet.title
# random.randint(A, B) - случайное целое число N, A ≤ N ≤ B.

barnaul_adr = []
volgograd_adr = []
voronej_adr = []
ekaterenburg_adr = []
izevsk_adr = []
irkuts_adr = []
kazan_adr = []
kaliningrad_adr = []
kemerovo_adr = []
krasnodar_adr = []
krasnoyars_adr = []
moskow_adr = []
naber_cheln_adr = []
nizniy_novgorod_adr = []
novosib_adr = []
omsk_adr = []
orenburg_adr = []
perm_adr = []
rostov_na_dony_adr = []
ryazan_adr = []
samara_adr = []
sankt_peter_adr = []
saratov_adr = []
sochi_and_adler = []
toliatty_adr = []
tomsk_adr = []
tyla_adr = []
tumen_adr = []
ufa_adr = []
chelabinsk_adr = []
yaroslavl_adr = []
musor_exela = []
#goroda_arr = [musor_exela, barnaul_adr, volgograd_adr, voronej_adr, ekaterenburg_adr, izevsk_adr, irkuts_adr, kazan_adr, kaliningrad_adr, kemerovo_adr, krasnodar_adr, krasnoyars_adr, moskow_adr, naber_cheln_adr, nizniy_novgorod_adr, novosib_adr, omsk_adr, orenburg_adr, perm_adr, rostov_na_dony_adr, ryazan_adr, samara_adr, sankt_peter_adr, saratov_adr, sochi_and_adler, toliatty_adr, tomsk_adr, tyla_adr, tumen_adr, ufa_adr, chelabinsk_adr, yaroslavl_adr]
goroda_arr = []
number_gorod = 0
#for cell in sheet[goroda_arr[number_gorod]]:
#    if cell.value == None:
#        break
#kolichestvo_gorodov = 28
kolichestvo_gorodov = 0

for cell in sheet['A']:
    kolichestvo_gorodov += 1
    print(cell.value)

#goroda_arr = []

print(kolichestvo_gorodov)

# СЧИТЫВАЕМ ГОРОДА иЗ ЭКСЕЛЯ ДИНАМИЧЕСКИ
for number_gorod in range(kolichestvo_gorodov):
    goroda_arr.append([])
    for cell in list(sheet.rows)[number_gorod]:
         if cell.value == None:
            number_gorod = number_gorod +1
            break
         else:
            goroda_arr[number_gorod].append(cell.value)

    # goroda_arr[number_adr].pop(0)

''''
massiv = []
for i in range(10):
    massiv.append([])
    for b in range(20):
         massiv[i].append(b)

print(massiv)
'''


goroda_arr.pop(0)
#print(goroda_arr)

print(goroda_arr)

vca = 'Программист'
slovar = {
    'Программист': 'txi'}
result = slovar.get(vca)
print(result)
#browser.find_elements_by_xpath(result)[0]




high_number_gorod = 0
number_goroda = 0
while number_goroda < len(goroda_arr):
    max_number = len(goroda_arr[number_goroda])
    if high_number_gorod< max_number:
        high_number_gorod = max_number
        number_goroda = number_goroda + 1
    else:
        number_goroda = number_goroda + 1
print('higt_number_gorog = ' + str(high_number_gorod))

