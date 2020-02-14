import time
import script3
from selenium import webdriver
from openpyxl import load_workbook
import random
import xlrd
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

wb = xlrd.open_workbook('DATA.xlsx')
sheet_xlr = wb.sheet_by_name('Вакансии')
a = 2
rows_sheet = sheet_xlr.nrows
print(rows_sheet)


wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)

sheet = wb[lst[0]]
sheet.title
iter_gorod = 0

# random.randint(A, B) - случайное целое число N, A ≤ N ≤ B.
while iter_gorod < len(script3.goroda_arr):
    iteracia = 0
    while iteracia < 5:

        i = random.randint(3, rows_sheet)
        ################### заводим содержимое в переменные ################################
        # print(a)
        vacancy = (sheet.cell(row=i, column=1).value)
        zp_ot = (sheet.cell(row=i, column=2).value)
        zp_do = (sheet.cell(row=i, column=3).value)
        opyt = (sheet.cell(row=i, column=4).value)
        obraz = (sheet.cell(row=i, column=5).value)
        pol = (sheet.cell(row=i, column=6).value)
        vozr_ot = (sheet.cell(row=i, column=7).value)
        vozr_do = (sheet.cell(row=i, column=8).value)
        opisanie = (sheet.cell(row=i, column=9).value)
        grafik = (sheet.cell(row=i, column=10).value)
        name_hr = (sheet.cell(row=i, column=11).value)
        mail = (sheet.cell(row=i, column=12).value)
        kod = (sheet.cell(row=i, column=13).value)
        phone = (sheet.cell(row=i, column=14).value)
        priem_zv_c = (sheet.cell(row=i, column=15).value)
        priem_zv_do = (sheet.cell(row=i, column=16).value)
        dni_priem = (sheet.cell(row=i, column=17).value)
        company = (sheet.cell(row=i, column=18).value)
        opis_company = (sheet.cell(row=i, column=19).value)
        # gorod = (sheet.cell(row=i, column=18).value)
        # print(priem_zv_c)

        ###############    autorith  ###############

        sheet = wb[lst[2]]
        sheet.title
        login = (sheet.cell(row=1, column=2).value)
        password = (sheet.cell(row=2, column=2).value)

        ###############    Open Browser and login  ###############


        #chrome_options.add_argument("--no-startup-window")
        browser = webdriver.Chrome()
        driver = browser
        browser.get('https://nn.rabota.ru/v3_myVacancy.html?action=create&company_registered=true&employer_registered=true')

        chrome_options = Options()
        chrome_options.add_argument("--no-startup-window")
        driver.refresh()
        driver.implicitly_wait(1)
        # Логинимся
        login_bar_xpath = '//*[@id="mail"]'
        login_bar = browser.find_element_by_xpath(login_bar_xpath)
        login_bar.send_keys(login)

        pass_bar_xpath = '//*[@id="password"]'
        pass_bar = browser.find_element_by_xpath(pass_bar_xpath)
        pass_bar.send_keys(password)
        # кликкаем войти
        driver.find_element_by_xpath('//*[@id="authForm"]/input[5]').click()


        #print(opisanie)


        ###################################################OLD##################################################
        time.sleep(1)

        vak_bar_xpath = '//*[@id="custom_position"]'
        vak_bar = browser.find_element_by_xpath(vak_bar_xpath)
        vak_bar.send_keys(vacancy)
        vak_bar.send_keys(Keys.ARROW_DOWN)
        vak_bar.send_keys(Keys.ENTER)
        time.sleep(3)

        driver.find_element_by_xpath('//*[@id="vacancyForm"]/div[1]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/span[1]').click()
        driver.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/label[5]/input[2]').send_keys(vacancy)
        driver.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/label[5]/input[3]').click()
        driver.find_element_by_xpath('//*[@id="jqmContent"]/div/div[3]/button').click()

        zp_ot_bar_xpath = '//*[@id="salary_from"]'
        zp_ot_bar = browser.find_element_by_xpath(zp_ot_bar_xpath)
        zp_ot_bar.send_keys(zp_ot)

        zp_do_bar_xpath = '//*[@id="salary_to"]'
        zp_do_bar = browser.find_element_by_xpath(zp_do_bar_xpath)
        zp_do_bar.send_keys(zp_do)

        driver.find_element_by_xpath('//*[@id="mce_3"]/button').click()

        ###############     Опыт работы     ###############

        #browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[1]/td[2]').click()
        if opyt == 'менее года':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[4]').click()
        elif opyt == '1 год':
            browser.find_element_by_xpath('//*[ @ id = "offer_experience_year_count"] / option[5]').click()
        elif opyt == '2 года':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[6]').click()
        elif opyt == '3 года':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[7]').click()
        elif opyt == '4 года':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[8]').click()
        elif opyt == '5 лет':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[9]').click()
        elif opyt == '6 лет':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[10]').click()
        elif opyt == '7 лет':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[11]').click()
        elif opyt == '8 лет':
             browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
        elif opyt == '9 лет':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
        elif opyt == '10 лет':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
        elif opyt == 'не имеет значения':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()
        elif opyt == 'без опыта':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[3]').click()

        ###############     Образование     ###############

        browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select').click()
        if obraz == "высшее":
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[2]').click()
        elif obraz == "неполное высшее":
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[3]').click()
        elif obraz == "среднее профессиональное":
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[4]').click()
        elif obraz == "среднее":
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[5]').click()
        elif obraz == "любое":
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[1]').click()

        ###############     Пол     ###############

        #browser.find_element_by_xpath('//*[@id="is_male"]').click()
        if pol == 'Не важно':
            browser.find_element_by_xpath('//*[@id="is_male"]/option[1]').click()
        elif pol == 'Мужской':
            browser.find_element_by_xpath('//*[@id="is_male"]/option[2]').click()
        elif pol == 'Женский':
            browser.find_element_by_xpath('//*[@id="is_male"]/option[3]').click()

        ###############     возраст     ###############

        vozr_ot_bar_xpath = '//*[@id="age_from"]'
        vozr_ot_bar = browser.find_elements_by_xpath(vozr_ot_bar_xpath)[0]
        vozr_ot_bar.send_keys(vozr_ot)

        vozr_do_bar_xpath = '//*[@id="age_to"]'
        vozr_do_bar = browser.find_elements_by_xpath(vozr_do_bar_xpath)[0]
        vozr_do_bar.send_keys(vozr_do)

        ###############    Контактная информация    ###############

        name_hr_bar_xpath = '//*[@id="vacancyForm"]/div[5]/div/div/table/tbody/tr[1]/td[2]/div/input'
        name_hr_bar = browser.find_element_by_xpath(name_hr_bar_xpath)
        name_hr_bar.clear()
        name_hr_bar.send_keys(name_hr)


        mail_bar_xpath = '//*[@id="contact_email"]'
        mail_bar = browser.find_element_by_xpath(mail_bar_xpath)
        mail_bar.clear()
        mail_bar.send_keys(mail)

        ###############    Phone    ###############

        kod_bar_xpath = '//*[@id="phoneContainerCode_"]'
        kod_bar = browser.find_elements_by_xpath(kod_bar_xpath)[0]
        kod_bar.clear()
        kod_bar.send_keys(kod)

        phone_bar_xpath = '//*[@id="phoneContainerNumber_"]'
        phone_bar = browser.find_elements_by_xpath(phone_bar_xpath)[0]
        phone_bar.clear()
        phone_bar.send_keys(phone)

        ###############    время для звонков    ###############

        browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]').click()

        if priem_zv_c == '0:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[1]').click()
        elif priem_zv_c == '0:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[2]').click()
        elif priem_zv_c == '1:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[3]').click()
        elif priem_zv_c == '1:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[4]').click()
        elif priem_zv_c == '2:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[5]').click()
        elif priem_zv_c == '2:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[6]').click()
        elif priem_zv_c == '3:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[7]').click()
        elif priem_zv_c == '3:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[8]').click()
        elif priem_zv_c == '4:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[9]').click()
        elif priem_zv_c == '4:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[10]').click()
        elif priem_zv_c == '5:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[11]').click()
        elif priem_zv_c == '5:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[12]').click()
        elif priem_zv_c == '6:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[13]').click()
        elif priem_zv_c == '6:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[14]').click()
        elif priem_zv_c == '7:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[15]').click()
        elif priem_zv_c == '7:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[16]').click()
        elif priem_zv_c == '8:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[17]').click()
        elif priem_zv_c == '8:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[18]').click()
        elif priem_zv_c == '9:00:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[19]').click()
        elif priem_zv_c == '9:30:00':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]/option[20]').click()
        elif priem_zv_c == '10:00:00':
             browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
        elif priem_zv_c == '10:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
        elif priem_zv_c == '11:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
        elif priem_zv_c == '11:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()
        elif priem_zv_c == '12:00:00':
            browser.find_element_by_xpath('// *[ @ id = "offer_experience_year_count"] / option[5]').click()
        elif priem_zv_c == '12:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[6]').click()
        elif priem_zv_c == '13:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[7]').click()
        elif priem_zv_c == '13:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[8]').click()
        elif priem_zv_c == '14:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[9]').click()
        elif priem_zv_c == '14:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[10]').click()
        elif priem_zv_c == '15:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[11]').click()
        elif priem_zv_c == '16:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
        elif priem_zv_c == '16:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
        elif priem_zv_c == '17:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
        elif priem_zv_c == '17:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()
        elif priem_zv_c == '18:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[4]').click()
        elif priem_zv_c == '18:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[5]').click()
        elif priem_zv_c == '19:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[6]').click()
        elif priem_zv_c == '19:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[7]').click()
        elif priem_zv_c == '20:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[8]').click()
        elif priem_zv_c == '20:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[9]').click()
        elif priem_zv_c == '21:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[10]').click()
        elif priem_zv_c == '21:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[11]').click()
        elif priem_zv_c == '22:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
        elif priem_zv_c == '22:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
        elif priem_zv_c == '23:00:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
        elif priem_zv_c == '23:30:00':
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()
        else:
            browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()

        ###############    дни для звонков    ###############
        browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]').click()
        if opyt == 'Рабочие дни':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[1]').click()
        elif opyt == 'Выходные дни':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[2]').click()
        elif opyt == 'Любой день':
            browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[3]').click()



        ###############    адрес работы    ###############

        browser.find_element_by_xpath('//*[@id="addressesList"]/div/a').click()
        # browser.find_element_by_xpath('//*[@id="vacancyAddressPopupLink"]').click()


        browser.find_element_by_xpath('//*[@id="vacancyCityPopupLink"]').click()
        browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[3]/div/p/a').click()
        browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[1]/input').send_keys(script3.goroda_arr[iter_gorod][0])
        time.sleep(2)
        browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[1]/input').send_keys(Keys.ARROW_DOWN + Keys.ENTER)
        #browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[3]/div/input').click()
        browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[3]/div/input').click()

        #browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[1]/input').sendKeys(Keys.ENTER)

        browser.find_element_by_xpath('//*[@id="vacancyAddressPopupLink"]').click()
        adr_bar_xpath = '//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[1]/td/input[1]'
        adr_bar = browser.find_element_by_xpath(adr_bar_xpath)
        adr_bar.clear()
        random_adres = random.randint(1, len(script3.goroda_arr[iter_gorod])-1)
        adr_bar.send_keys(script3.goroda_arr[iter_gorod][0] + ' ' + script3.goroda_arr[iter_gorod][random_adres])
        time.sleep(1)
        adr_bar.send_keys(Keys.ARROW_DOWN)
        adr_bar.send_keys(Keys.ENTER)
        time.sleep(1)
        browser.find_element_by_xpath('//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[1]/td/input[1]').click()
        time.sleep(1)
        browser.find_element_by_xpath('//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[4]/td/input').click()
        time.sleep(1)
        browser.find_element_by_xpath('//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[6]/td/input').click()

        
       # browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div/div[2]/div/div/table/tbody/tr/td[1]/div/div[2]/div[2]/div[2]/div/a').click()
       # browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div/div[3]/div/input').click()

        #browser.find_element_by_xpath('//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[6]/td/input').click()
        #adr_bar.send_keys(Keys.ENTER)
        ###############  anonim company  ###############

        browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[2]/tbody/tr[3]/td[2]/div/label/input').click()

        browser.find_element_by_xpath('/html/body/div[2]/div[2]/div[2]/div/div[3]/div[2]/div/div/form/div[6]/div/div/table[2]/tbody/tr[3]/td[2]/span/input').send_keys('Курьерская служба доставки')
        ###############    Выбор рубрики    ###############

        browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[1]/tbody/tr/td[2]/div[1]/a').click()
        rubriks = script3.xpath_rubriks.get(company, 0)
        print(vacancy)
        print(company)
        print(rubriks)
        browser.find_element_by_xpath(rubriks).click()

        pod_rubriks = script3.vibor_podrubrik.get(opis_company, 0)
        print(pod_rubriks)
        xpartt = script3.pod_rubriks.get(vacancy) #тестовая строка
        print(xpartt)
      # browser.find_element_by_xpath(xpartt).click()
        #print(xpartt)#тестовая
      # browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input').click()
      # browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input').click()
        time.sleep(1)
        browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/div[3]/a[1]').click()

        ###############    График работы    ###############

        if grafik == 'полный рабочий день':
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[1]/label/input').click()
        elif grafik =='сменный график':
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[2]/label/input').click()
        elif grafik == 'свободный график':
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[3]/label/input').click()
        elif grafik == 'частичная занятость':
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[4]/label/input').click()
        elif grafik == 'удаленная работа':
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[5]/label/input').click()
        elif grafik == 'вахта':
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[6]/label/input').click()

        ####################################################


        iframe = browser.find_elements_by_xpath('//*[@id="description_ifr"]')[0]
        driver.switch_to.default_content()
        driver.switch_to.frame(iframe)
        div = browser.find_elements_by_tag_name('p')[0]

        div.send_keys(opisanie)
        driver.switch_to.default_content()

        browser.find_element_by_xpath('//*[@id="publishButton"]').click()
        #browser.find_element_by_xpath('//*[@id="publishButton"]').click()
        browser.quit()
        iteracia = iteracia + 1
        print(iteracia)
    iter_gorod = iter_gorod + 1
    print(iter_gorod)
