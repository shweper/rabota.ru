from openpyxl import load_workbook
import random

wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)
list1 = lst[0]
# print(random.choice(lst))
# Берём рандомный лист
sheet = wb[list1]
sheet.title
# заводим содержимое в переменные
# random.randint(A B) - случайное целое число N, A ≤ N ≤ B.
a = 3
vacancy = (sheet.cell(row=a, column=1).value)
for row in sheet.iter_cols(max_col=1):
    for value in row:
            a = a+1



print(a)
