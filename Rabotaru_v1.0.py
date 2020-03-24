import time

from selenium.webdriver import ActionChains

import script3
from selenium import webdriver
from openpyxl import load_workbook
import random
import xlrd
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

#Для оптимизации
optimiz = False
popitka = 0

#Переменные для логирования
error_string = 0
error_gorods = []
vsego = 0
while vsego <= 1500:

    #Рубрики
    xpath_rubriks = {
        'IT / Интернет / Телеком': '//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[1]/li/a',
        'Топ-менеджмент': '//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[2]/li/a',

        'Банки / Инвестиции / Ценные бумаги': '//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/a',
        'Бухгалтерия / Аудит / Экономика предприятия':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/a',
        'Страхование':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[3]/a',

        'HR / Кадры / Подбор персонала':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[1]/a',
        'Административный персонал':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[2]/a',
        'Консалтинг / Тренинги':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[3]/a',
        'Охрана / Безопасность':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[4]/a',
        'Юриспруденция':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[5]/a',
        'Дизайн / Полиграфия':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[6]/a',
        'Маркетинг / Реклама / PR':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[7]/a',
        'СМИ / Издательства':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[8]/a',

        'Госслужба / Некоммерческие организации':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[6]/li/ul/li[1]/a',
        'Культура / Искусство / Развлечения':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[6]/li/ul/li[2]/a',
        'Образование / Наука':'//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[6]/li/ul/li[3]/a',

        'Торговля':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[1]/li/a',
        'Производство / Агропром':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[2]/li/a',

        'Недвижимость / Риелторские услуги':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/a',
        'Строительство / ЖКХ / Эксплуатация':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/a',

        'Логистика / Склад / ВЭД':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[4]/li/ul/li[1]/a',
        'Транспорт / Автобизнес / Автосервис':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[4]/li/ul/li[2]/a',

        'Красота / Фитнес / Спорт':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[5]/li/ul/li[1]/a',
        'Медицина / Фармация / Ветеринария':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[5]/li/ul/li[2]/a',

        'Бытовые услуги / Обслуживание оборудования':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[1]/a',
        'Домашний персонал':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[2]/a',
        'Рестораны / Питание':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[3]/a',
        'Туризм / Гостиницы':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[4]/a',
        'Работа без специальной подготовки / Без опыта':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[7]/li/a',
        'Работа для студентов / Стажировки':'//*[@id="jqmContent"]/div/div[1]/ul/li[2]/ul[8]/li/a',

    }
    #Словари для рубрик
    #IT / Интернет / Телеком
    xpath_it_ithernet = {
        #Программирование, разработка
        'Ведущий разработчик' : '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Программист' : '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Проектировщик': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Тестировщик': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        #
        #Виды ПО
        #
        '1С': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'ERP-системы': '//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Банковское ПО': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Игровое ПО': '//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',
        #
        #Аналитика
        #
        'Бизнес-аналитик': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/label/input',
        'Системный аналитик': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/label/input',
        'Технический писатель': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[3]/label/input',
        #
        #Сети, защита информации
        #
        'Компьютерная, информационная безопасность': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[1]/label/input',
        'Системное администрирование, DBA': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[2]/label/input',
        #
        #Монтаж, техобслуживание
        #
        'Беспроводные технологии, сотовая связь': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[5]/li/ul/li[1]/label/input',
        'Монтаж оборудования и сетей': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[5]/li/ul/li[2]/label/input',
        'Обслуживание банкоматов и платежных терминалов': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[5]/li/ul/li[3]/label/input',
        'Сборка ПК': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[5]/li/ul/li[4]/label/input',
        'Техподдержка, сервисное обслуживание': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[5]/li/ul/li[5]/label/input',
        #
        #Продажи, работа с клиентами
        #
        'Продажи, работа с клиентами': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Тендеры': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Электронная коммерция': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        #
        #Маркетинг, консалтинг
        #
        'IT-консалтинг': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Интернет-маркетинг, SMO, SMM': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Поисковая оптимизация, SEO': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        #
        #Дизайн, верстка, контент
        #
        'Веб-дизайн': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Верстка':  '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Контент, модерация': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Мультимедиа': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input',
        #
        # Операторы
        #
        'Call-центр': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[4]/li/ul/li[1]/label/input',
        'ПК, БД': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[4]/li/ul/li[2]/label/input',
        #
        #Руководство
        #
        'IT-директор, технический директор': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[5]/li/ul/li[1]/label/input',
        'Директор, управляющий': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[5]/li/ul/li[2]/label/input',
        'Руководитель отдела, заместитель руководителя': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[5]/li/ul/li[2]/label/input',
        'Руководитель проекта': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[5]/li/ul/li[4]/label/input',
        'Небольшой опыт, нет опыта': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[6]/li/label/input',
        'Другое': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[7]/li/label/input'

    }
    #Топ-менеджмент
    xpath_top_menegment = {
        #Административное управление
        #
        'HR-директор, директор по персоналу': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Директор, управляющий': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Директор филиала, офиса': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Исполнительный директор, административный директор': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        #
        #Функциональное управление
        #
        'IT-директор, технический директор': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Арт-директор, креативный директор': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        #
        #Финансы, коммерция, развитие
        #
        'Директор по маркетингу': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Директор по развитию': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Коммерческий директор, директор по продажам': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Финансовый директор': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Другое': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/label/input',
    }

    #ФИНАНСЫ / СТРАХОВАНИЕ
    #Банки / Инвестиции / Ценные бумаги
    xpath_banki_invest = {
        #Операции, направления
        #
        'Валютные операции': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Лизинг': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Межбанковские операции': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Налогообложение':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Обслуживание банкоматов': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Пластиковые карты': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',
        'Финансовая аналитика': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[7]/label/input',
        'Ценные бумаги': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[8]/label/input',
        #
        #Работа с клиентами
        #
        'Консультирование, работа с клиентами': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Кредитование': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Продажи финансовых услуг': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Расчетно-кассовое обслуживание': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        #
        #Руководство
        #
        'Руководитель департамента, высшее руководство': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Руководитель отдела, заместитель руководителя': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Небольшой опыт, нет опыта': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/label/input',
        'Другое': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[4]/li/label/input'

    }
    #Бухгалтерия / Аудит / Экономика предприятия
    xparh_buhgalteria = {
        #Бухгалтерия
        #
        '1C': '//*[@id="jqmContent"]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Бухгалтер': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Казначейство': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Касса': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'МСФО, GAAP':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Налоговый учет': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',
        'Первичная документация': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[7]/label/input',
        'Расчет заработной платы': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[8]/label/input',
        'Расчет себестоимости': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[9]/label/input',
        'Расчеты с покупателями': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[10]/label/input',
        'Составление отчетности': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[11]/label/input',
        'Учет основных средств и материальных активов': '/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[12]/label/input',
        #
        #Аудит, экономика предприятия
        #
        'Аудит, внутренний контроль': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Планово-экономическое управление': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Финансовый консалтинг': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        #
        #Руководство
        #
        'Главный бухгалтер': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Заместитель главного бухгалтера': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Руководитель отдела, заместитель руководителя': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Финансовый директор': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Небольшой опыт, нет опыта': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/label/input',
        'Другое': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[4]/li/label/input'
    }

    #Страхование
    xpath_strahovanie= {
        #Имущественное страхование
        #
        'Автострахование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Страхование недвижимости, имущества':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        #
        #Личное страхование
        #
        'Медицинское страхование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Страхование жизни':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',

    #Страхование бизнеса

        'Корпоративное страхование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li/label/input',

        #Оценка, урегулирование убытков

        'Оценка, экспертиза':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Урегулирование убытков':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',

        #Руководство

        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор филиала':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Руководитель отдела, заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[4]/li/label/input'
    }


    #ОФИСНЫЕ СЛУЖБЫ / БИЗНЕС-УСЛУГИ
    #HR / Кадры / Подбор персонала
    xpath_hr_kadri = {
        #Кадровое дело

        'Кадровое делопроизводство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Охрана труда':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',

    #Развитие персонала

        'Льготы и компенсации':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Обучение и развитие персонала':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Тренинги':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',

        #Консультирование

        'Кадровый консалтинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/label/input',
        'Психологические консультации':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/label/input',

    #Подбор персонала

        'Аутстаффинг, лизинг персонала':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Подбор домашнего персонала':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Подбор сотрудников, рекрутмент':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Региональный подбор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Хедхантинг, executive search':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',

    #Руководство

        'HR-директор, директор по персоналу':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Руководитель отдела, заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input'
    }

    #Административный персонал
    xpath_adm_personal= {
        #Секретариат, делопроизводство

        'Архив':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Делопроизводство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Офис-менеджер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Переводчик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Помощник руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'Ресепшн, приемная':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',
        'Секретарь':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[7]/label/input',

        #Операторы

        'Оператор call-центра, на телефоне':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Оператор ПК, БД':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',

        #АХО

        'Администратор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Водитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Диспетчер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Курьер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Специалист АХО':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',
        'Уборщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[6]/label/input',

        #Руководство

        'Исполнительный директор, административный директор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Руководитель отдела, заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input'

    }
    #Консалтинг / Тренинги
    xpath_konsalting = {
        #Консалтинг
        'IT-консалтинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Кадровый консалтинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Консалтинг в области налогов и бухучета':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Управленческий консалтинг, организационное развитие':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Юридический консалтинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',

        #Исследования
        'Аналитика, исследования рынка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li/label/input',

        #Продажи
        'Продажа услуг, работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li/label/input',

        #Тренинги
        'Психологическое консультирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Тренинги, коучинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        #Руководство
        'Директор по развитию':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Руководитель отдела, заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input'
        }

    #Охрана / Безопасность
    xpath_ohrana= {
        'Инкассация':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Личная охрана, охрана мероприятий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Охрана объектов, сопровождение грузов':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Сторож / Вахтер / Консьерж':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',

        'Коллекторская деятельность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Экономическая безопасность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',

        'Сотрудники полиции, МВД, МЧС':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li/label/input',

        'Системы безопасности и видеонаблюдения':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li/label/input',

        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Начальник охраны, руководитель СБ':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',

        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input'
    }

    #Юриспруденция
    xpath_urisprud= {
        'Авторское право, интеллектуальная собственность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Административное право':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Гражданское право':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Трудовое право':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Уголовное право':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Финансовое, налоговое право':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',

        'Адвокатура, судебная деятельность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Арбитраж':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Лицензирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Нотариат':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Регистрация, ликвидация предприятий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',
        'Юридическое консультирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[6]/label/input',
        'Юридическое сопровождение бизнеса':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[7]/label/input',

        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Руководитель отдела, заместитель руководителя:':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input'
    }


    #МАРКЕТИНГ / РЕКЛАМА / СМИ
    #Дизайн / Полиграфия
    xpath_design= {
        '3D-дизайн, промышленный дизайн':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Веб-дизайн, мультимедиа':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Дизайнер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Дизайн интерьера':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Дизайн одежды, аксессуаров':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Дизайн печатных изданий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',
        'Дизайн упаковки, рекламной и сувенирной продукции':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[7]/label/input',
        'Ландшафтный дизайн':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[8]/label/input',
        'Флористика':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[9]/label/input',

        'Продажа полиграфических услуг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li/label/input',

        'Препресс, верстка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Технолог печати, инженер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Типографские мастера и рабочие':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',

        'Арт-директор, креативный директор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Руководитель отдела, заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[5]/label/input'
    }
    #Маркетинг / Реклама / PR
    xpath_marketing= {
        'BTL-услуги':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Аналитика, исследования рынка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Бренд-менеджмент':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Интернет-маркетинг, SMO, SMM':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Консультирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'Мерчандайзинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',
        'Рекламно-сувенирная продукция':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[7]/label/input',
        'Трейд-маркетинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[8]/label/input',

        'Event-услуги':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Выставочная деятельность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Копирайтер, редактор, журналист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Маркетинговые коммуникации, PR':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',
        'Промоутер, модель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[5]/label/input',

        'Медиапланирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Продажа рекламных услуг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Производство рекламы':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',

        'Директор по маркетингу, развитию':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор по продажам':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Креативный директор, арт-директор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Руководитель проекта, направления':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[6]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[7]/label/input'

    }

    #СМИ / Издательства
    xpath_smi= {
        'Журналистика':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Корректура':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Перевод':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Редакторское дело':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',

        'Дизайн, верстка, препресс':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Дистрибуция, распространение печатной продукции':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Издательская деятельность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Фотосъемка, фотодело':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',

        'Продажа рекламных услуг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',

        'Производство телепрограмм и кинофильмов':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Радио':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Телевидение':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',

        'Арт-директор, креативный директор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Главный редактор, шеф-редактор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Коммерческий директор, директор по продажам':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Руководитель отдела, заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[6]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[7]/label/input',
    }


    #КУЛЬТУРА / ОБРАЗОВАНИЕ / ГОССЛУЖБА
    #Госслужба / Некоммерческие организации
    xpath_gossluzba = {
        'Лицензирование, экспертиза':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Социальная защита':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Управление, контроль':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',

        'Благотворительность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Волонтерство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Общественная деятельность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',

        'Директор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Руководитель отдела, заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input'

    }

    #Культура / Искусство / Развлечения
    xpath_kultura= {
        'Изобразительное искусство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Искусствоведение , реставрация, антиквариат':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Мода':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Фотография':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',

        'Актер , артист эстрады':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Кино, мультипликация':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Музыка, вокальное искусство:':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Театр':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',
        'Хореография, балет':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[5]/label/input',

        'Библиотечное дело, архивное дело':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Литература, переводы':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',

        'Организация мероприятий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Технический, административный персонал':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',

        'Директор, управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Руководитель отдела, коллектива':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input',
    }

    #Образование / Наука
    xpath_obrazovanie= {
        'Воспитатель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Преподаватель вуза, учебного центра':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Преподаватель иностранного языка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Репетитор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Учитель начальных классов':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'Учитель среднего и старшего звена':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',

        'Консалтинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Логопедия, дефектология':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Психология':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Тренинги':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',

        'Научная деятельность, лабораторные работы':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li/label/input',

        'Методист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li/label/input',

        'Директор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Руководитель отдела, направления':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Небольшой опыт, нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input'

    }

    #Торговля
    xpath_torgovla= {
        'Администратор магазина, торгового зала':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Менеджер по продажам':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Мерчандайзер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Продавец / Кассир':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Продавец-консультант':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'Продажи по телефону / Телемаркетинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',
        'Товаровед / Учет товаров':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[7]/label/input',
        'Торговый представитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[8]/label/input',

        'Менеджер по закупкам':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Снабжение':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',

        'Продажи по каталогам / MLM':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li/label/input',

        'Продажи по телефону / Телемаркетинг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Продажи / Работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Работа с регионами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Торговый представитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',

        'Директор по развитию':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Коммерческий директор / Директор по продажам':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Руководитель отдела / Заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[5]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[6]/label/input'
    }

    #Производство / Агропром
    xpath_proizvodstvo= {
        'Инженер / Технолог / Конструктор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'КИПиА / Метрология':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Контроль качества':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Охрана труда / Экология':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Проектирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'Сертификация':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',

        'Монтажник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Оператор станков / Автоматических линий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Сборщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Сварщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',
        'Столяр':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[5]/label/input',
        'Техник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[6]/label/input',
        'Токарь':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[7]/label/input',
        'Фрезеровщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[8]/label/input',
        'Швея':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[9]/label/input',
        'Электрик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[10]/label/input',
        'Электромонтажник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[11]/label/input',
        'Ювелир':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[12]/label/input',

        'Продажи / Работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/label/input',
        'Снабжение / Закупки / Тендеры':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/label/input',

        'Автомобилестроение, производство автозапчастей':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Атомная и другие виды энергетики':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Деревообработка, производство мебели':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Добывающая промышленность':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Животноводство / Мясопереработка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',
        'Легкая промышленность, производство ТНП':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[6]/label/input',
        'Металлургия, металлообработка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[7]/label/input',
        'Нефтегазовая отрасль':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[8]/label/input',
        'Оборонное производство, авиакосмическая отрасль':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[9]/label/input',
        'Пищевое производство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[10]/label/input',
        'Приборостроение, радиотехника, электроника':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[11]/label/input',
        'Производство пищевого и торгового оборудования':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[12]/label/input',
        'Производство стройматериалов и конструкций':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[13]/label/input',
        'Растениеводство / Сельхозпроизводство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[14]/label/input',
        'Станкостроение, производство техники и оборудования':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[15]/label/input',
        'Судостроение, судоремонт':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[16]/label/input',
        'Фармацевтическое производство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[17]/label/input',
        'Химия, нефтехимия':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[18]/label/input',
        'Другие виды производства':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[19]/label/input',

        'Главный инженер / Главный специалист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор по развитию':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Коммерческий директор / Директор по продажам':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Мастер участка / смены / Бригадир':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input',
        'Начальник производства / цеха / отдела':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[6]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[7]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[8]/label/input'
    }


    #СТРОИТЕЛЬСТВО / НЕДВИЖИМОСТЬ
    #Недвижимость / Риелторские услуги
    xpath_nedvizimost= {
        #Жилая недвижимость
        'Аренда':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Продажа':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        #Коммерческая недвижимость (временно поставлю 1, чтоб не дублировать ключи
        'Аренда1':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Продажа1':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',

        'Консалтинг / Оценка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li/label/input',

        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор филиала':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Руководитель отдела / направления':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input'

    }

    #Строительство / ЖКХ / Эксплуатация
    xpath_stroitelstvo_zkh= {
        'Архитектура':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Геодезия / Землеустройство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Замерщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Инженер ПТО':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Проектирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'Сметное дело':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',
        'Снабжение / Закупки / Тендеры':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[7]/label/input',

        'Вентиляция / Климатическое оборудование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Водоснабжение / Водоотведение / Отопление':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Газоснабжение':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Инженер по сервису':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',
        'Инженер по сетям':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[5]/label/input',
        'Инженер-строитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[6]/label/input',
        'КИПиА / Метрология':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[7]/label/input',
        'Наружные трубопроводы и коммуникации':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[8]/label/input',
        'Электроснабжение / Слаботочные системы':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[9]/label/input',
        'Энергетик / Теплотехник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[10]/label/input',

        'Благоустройство территорий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/label/input',
        'Лифты':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/label/input',
        'Машинист/водитель спецтехники':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[3]/label/input',
        'Эксплуатация зданий / Коммунальные службы':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[4]/label/input',

        'Охрана труда / Техника безопасности':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[1]/label/input',
        'Технадзор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[4]/li/ul/li[2]/label/input',

        'Бетонщик / Арматурщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Дорожный рабочий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Каменщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Кровельщик / Жестянщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Маляр':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',
        'Машинист стройтехники':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[6]/label/input',
        'Монтажник / Сборщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[7]/label/input',
        'Монтажник-слаботочник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[8]/label/input',
        'Отделочник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[9]/label/input',
        'Плотник / Столяр':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[10]/label/input',
        'Промышленный альпинист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[11]/label/input',
        'Сантехник / Слесарь-сантехник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[12]/label/input',
        'Сварщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[13]/label/input',
        'Слесарь / Сборщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[14]/label/input',
        'Штукатур / Плиточник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[15]/label/input',
        'Электрик / Электромонтер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[16]/label/input',
        'Электромонтажник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[17]/label/input',

        'Главный инженер / Ведущий специалист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Коммерческий директор / Директор по продажам':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Прораб':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Руководитель отдела / участка / Бригадир':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[6]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[7]/label/input'
    }

    #ТРАНСПОРТ / ЛОГИСТИКА
    #Логистика / Склад / ВЭД
    xpath_logistika = {
        'Водитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Логистика':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Экспедирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',

        'Менеджер по ВЭД':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Таможенное оформление':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',

        'Закупки / Снабжение':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/label/input',
        'Продажа логистических услуг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/label/input',

        'Грузчик / Комплектовщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Кладовщик / Товаровед':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Оператор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Упаковщик / Фасовщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',

        'Административный персонал':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Диспетчер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',

        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Заведующий / начальник склада':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Руководитель отдела / Заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[5]/label/input'
    }

    #Транспорт / Автобизнес / Автосервис
    xpath_transport= {
        'Автобус / маршрутное такси':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Газель (малый коммерческий а/т)':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'На служебном грузовом а/м':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'На служебном легковом а/м':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'С личным грузовым а/м':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'С личным легковым а/м':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',
        'Спецтехника':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[7]/label/input',
        'Такси':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[8]/label/input',
        'Троллейбус / трамвай':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[9]/label/input',
        'Экспедирование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[10]/label/input',

        'Лизинг / прокат автотранспорта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Продажа автозапчастей и автохимии':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Продажа коммерческого транспорта / спецтехники':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',
        'Продажа легковых автомобилей':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[4]/label/input',
        'Продажа мототехники и запчастей':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[5]/label/input',

        'Авиация':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/label/input',
        'Водный транспорт':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/label/input',
        'Железнодорожный транспорт':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[3]/label/input',
        'Метрополитен':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[4]/label/input',

        'Автомойщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Администратор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Диспетчер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Заправщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',

        'Жестянщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Маляр / Колорист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Слесарь / Механик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Электрик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Арматурщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input',
        'Инженер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[6]/label/input',
        'Установщик дополнительного оборудования':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[7]/label/input',
        'Шиномонтажник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[8]/label/input',

        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Руководитель отдела / Заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[4]/label/input'
    }

    #КРАСОТА / ЗДОРОВЬЕ
    #Красота / Фитнес / Спорт
    xpath_krasota= {
        'Визажист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Косметология':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Маникюр / Педикюр':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Массаж':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Парикмахер / Стилист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Тренер / Инструктор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',

        'Продажи услуг / Работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li/label/input',

        'Администратор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input'
    }

    #Медицина / Фармация / Ветеринария
    xpath_medicina= {
        'Аллерголог-иммунолог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Анестезиолог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Гастроэнтеролог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Гинеколог / Акушер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Кардиолог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Косметолог / Дерматовенеролог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',
        'Медицинская диагностика / УЗИ / Рентген':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[7]/label/input',
        'Невролог / Мануальный терапевт':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[8]/label/input',
        'Онколог / Маммолог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[9]/label/input',
        'Отоларинголог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[10]/label/input',
        'Офтальмолог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[11]/label/input',
        'Педиатр':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[12]/label/input',
        'Психиатр / Нарколог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[13]/label/input',
        'Стоматолог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[14]/label/input',
        'Терапевт':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[15]/label/input',
        'Травматолог / Ортопед':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[16]/label/input',
        'Уролог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[17]/label/input',
        'Физиотерапевт / ЛФК':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[18]/label/input',
        'Хирург':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[19]/label/input',
        'Эндокринолог':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[20]/label/input',
        'Другие специалисты':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[21]/label/input',

        'Ассистент':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Лаборант':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Массажист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Медсестра':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Санитар':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',
        'Сиделка / Няня':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[6]/label/input',
        'Фельдшер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[7]/label/input',

        'Обслуживание медтехники и оборудования':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Оптика':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Продажа медицинских товаров / услуг':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Фармацевт / Провизор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',

        'Клинические испытания / Лабораторные исследования':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Консультирование / Экспертиза':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',
        'Медицинский представитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[3]/label/input',

        'Ветеринария':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[4]/li/ul/li/label/input',

        'Регистратура / Административный персонал':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[5]/li/ul/li/label/input',

        'Главврач / Заведующий отделением':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[1]/label/input',
        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[2]/label/input',
        'Руководитель отдела / Заместитель руководителя':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[3]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[6]/li/ul/li[5]/label/input'

    }

    #СФЕРА УСЛУГ
    #Бытовые услуги / Обслуживание оборудования
    xpath_bitovie_yslugy= {
        'Металлоремонт':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Пошив и ремонт одежды':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Ремонт часов, ювелирных изделий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Ремонт электроники, ПК, телефонов':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Уборка помещений / Клининговые услуги':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',
        'Фото / видеосъемка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[6]/label/input',
        'Химчистка / Прачечная':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[7]/label/input',

        'Прием заказов / Диспетчер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Продажа услуг / Работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',

        'Пищевое и холодильное оборудование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Промышленное оборудование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Системы вентиляции и климатическая техника':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Строительная и спецтехника':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Торговое и складское оборудование':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',

        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Руководитель отдела / службы':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input'
    }

    #Домашний персонал
    xpath_home_personal= {
        'Водитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Горничная':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Домработница':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Повар':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Помощник по хозяйству':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Разнорабочий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',
        'Садовник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[7]/label/input',
        'Семейная пара':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[8]/label/input',
        'Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[9]/label/input',
        'Экономка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[10]/label/input',

        'Воспитатель / Гувернантка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Няня':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Сиделка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input'

    }

    #Рестораны / Питание
    xpath_restorani= {
        'Кондитер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Пекарь':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Пиццамейкер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Повар':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Посудомойщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',
        'Сушист':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[6]/label/input',

        'Бармен / Бариста':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Буфетчик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Кассир':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Официант':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Сомелье':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[5]/label/input',
        'Хостес / Администратор зала':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[6]/label/input',
        'Кальянщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[7]/label/input',

        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Су-шеф':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Шеф-повар / Заведующий производством':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[5]/label/input'
    }

    #Туризм / Гостиницы
    xpath_turizm= {
        'Авиа / железнодорожные билеты':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Бронирование гостиниц':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[2]/label/input',
        'Визовая поддержка':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[3]/label/input',
        'Продажи / Работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[4]/label/input',
        'Экскурсионное обслуживание':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[5]/label/input',

        'Администратор / Портье':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Обслуживающий персонал':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Организация и проведение мероприятий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Продажи / Работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',

        'Директор / Управляющий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Руководитель отдела / направления':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Небольшой опыт / Нет опыта':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[4]/label/input'

    }

    #Работа без специальной подготовки / Без опыта
    xpath_job_no_opit= {
        'Агент':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[1]/label/input',
        'Мерчандайзинг / Выкладка товара':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[2]/label/input',
        'Продажи / Работа с клиентами':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[3]/label/input',
        'Тайный покупатель / Тайный клиент':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[4]/label/input',
        'Торговый представитель':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[1]/li/ul/li[5]/label/input',

        'Грузчик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[1]/label/input',
        'Разнорабочий':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[2]/label/input',
        'Складские работы1':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[2]/li/ul/li[3]/label/input',

        'Гардеробщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[1]/label/input',
        'Дворник':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[2]/label/input',
        'Заправщик / Парковщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[3]/label/input',
        'Персонал гостиниц и ресторанов':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[4]/label/input',
        'Прием заказов':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[5]/label/input',
        'Сторож / Вахтер / Консьерж':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[6]/label/input',
        'Уборщик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul[3]/li/ul/li[7]/label/input',

        'Оператор call-центра / на телефоне':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Оператор ПК':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',

        'Администратор':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[1]/label/input',
        'Диспетчер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[2]/label/input',
        'Курьер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[2]/li/ul/li[3]/label/input',

        'Промоутер / Проведение опросов':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[1]/label/input',
        'Расклейщик / Раздатчик':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/ul/li[2]/label/input',

        'Домашний персонал':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[4]/li/ul/li/label/input',
        'Другое':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[5]/li/label/input'

    }
    #Работа для студентов / Стажировки
    xpath_job_for_student= {
        'IT / Интернет / Телеком':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Маркетинг / Реклама / СМИ':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Продажи':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Производство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Рестораны / Гостиницы / Туризм':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Строительство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Финансы / Банки':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',
        'Юриспруденция':'/html/body/div[6]/div[1]/div/div[1]/ul/li[1]/ul/li/ul/li[1]/label/input',

        'Call-центр':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[1]/label/input',
        'Делопроизводство':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[2]/label/input',
        'Курьер':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[3]/label/input',
        'Оператор ПК':'/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[1]/li/ul/li[4]/label/input',
        'Стажировка': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/label/input',
        'Другое': '/html/body/div[6]/div[1]/div/div[1]/ul/li[2]/ul[3]/li/label/input',

    }

    iter_gorod = 0
    vibor_podrubrik = {
    'IT / Интернет / Телеком': script3.xpath_it_ithernet,
    'Топ-менеджмент': script3.xpath_top_menegment,

    #'ФИНАНСЫ / СТРАХОВАНИЕ':

    'Банки / Инвестиции / Ценные бумаги': script3.xpath_banki_invest,
    'Бухгалтерия / Аудит / Экономика предприятия': script3.xparh_buhgalteria,
    'Страхование':script3.xpath_strahovanie,

    #'ОФИСНЫЕ СЛУЖБЫ / БИЗНЕС-УСЛУГИ'

    'HR / Кадры / Подбор персонала':script3.xpath_hr_kadri,
    'Административный персонал': script3.xpath_adm_personal,
    'Консалтинг / Тренинги':script3.xpath_konsalting,
    'Охрана / Безопасность':script3.xpath_ohrana,
    'Юриспруденция':script3.xpath_urisprud,

    #'МАРКЕТИНГ / РЕКЛАМА / СМИ'

    'Дизайн / Полиграфия':script3.xpath_design,
    'Маркетинг / Реклама / PR':script3.xpath_marketing,
    'СМИ / Издательства':script3.xpath_smi,

    #'КУЛЬТУРА / ОБРАЗОВАНИЕ / ГОССЛУЖБА'

    'Госслужба / Некоммерческие организации':script3.xpath_gossluzba,
    'Культура / Искусство / Развлечения':script3.xpath_kultura,
    'Образование / Наука':script3.xpath_obrazovanie,
    'Торговля':script3.xpath_torgovla,
    'Производство / Агропром':script3.xpath_proizvodstvo,

    #'СТРОИТЕЛЬСТВО / НЕДВИЖИМОСТЬ'

    'Недвижимость / Риелторские услуги':script3.xpath_nedvizimost,
    'Строительство / ЖКХ / Эксплуатация':script3.xpath_stroitelstvo_zkh,

    #'ТРАНСПОРТ / ЛОГИСТИКА'

    'Логистика / Склад / ВЭД':script3.xpath_logistika,
    'Транспорт / Автобизнес / Автосервис':script3.xpath_transport,

    #'КРАСОТА / ЗДОРОВЬЕ'

    'Красота / Фитнес / Спорт':script3.xpath_krasota,
    'Медицина / Фармация / Ветеринария':script3.xpath_medicina,

    #'СФЕРА УСЛУГ'

    'Бытовые услуги / Обслуживание оборудования':script3.xpath_bitovie_yslugy,
    'Домашний персонал':script3.xpath_home_personal,
    'Рестораны / Питание':script3.xpath_restorani,
    'Туризм / Гостиницы':script3.xpath_turizm,
    'Работа без специальной подготовки / Без опыта':script3.xpath_job_no_opit,
    'Работа для студентов / Стажировки':script3.xpath_job_for_student
    }


    iteracia = 1
    while iteracia < script3.high_number_gorod:
        iter_gorod = 0
        while iteracia < len(script3.goroda_arr[iter_gorod]):
          #  try:
            print(len(script3.goroda_arr[iter_gorod]))
            error_string = 1043
            wb = xlrd.open_workbook('./DATA.xlsx')
            sheet_xlr = wb.sheet_by_name('Вакансии')
            a = 2
            rows_sheet = sheet_xlr.nrows
            #print(rows_sheet)


            wb = load_workbook('./DATA.xlsx')
            lst = (wb.sheetnames)

            sheet = wb[lst[0]]
            sheet.title
            error_string = 1056

            # random.randint(A, B) - случайное целое число N, A ≤ N ≤ B.

            error_string = 1060
            i = random.randint(3, rows_sheet)
            ################### заводим содержимое в переменные ################################
            # print(a)
            vacancy = (sheet.cell(row=i, column=1).value)
            zp_ot = (sheet.cell(row=i, column=2).value)
            zp_do = (sheet.cell(row=i, column=3).value)
            opyt = (sheet.cell(row=i, column=4).value)
            obraz = (sheet.cell(row=i, column=5).value)
            pol = (sheet.cell(row=i, column=6).value)
            vozr_ot = (sheet.cell(row=i, column=7).value)
            vozr_do = (sheet.cell(row=i, column=8).value)
            opisanie = (sheet.cell(row=i, column=9).value)
            grafik = (sheet.cell(row=i, column=10).value)
            name_hr = (sheet.cell(row=i, column=11).value)
            mail = (sheet.cell(row=i, column=12).value)
            kod = (sheet.cell(row=i, column=13).value)
            phone = (sheet.cell(row=i, column=14).value)
            koment = (sheet.cell(row=i, column=15).value)
            priem_zv_c = (sheet.cell(row=i, column=16).value)
            priem_zv_do = (sheet.cell(row=i, column=17).value)
            dni_priem = (sheet.cell(row=i, column=18).value)
            company = (sheet.cell(row=i, column=19).value)
            opis_company = (sheet.cell(row=i, column=20).value)
            company1 = (sheet.cell(row=i, column=21).value)
            opis_company1 = (sheet.cell(row=i, column=22).value)
            company2 = (sheet.cell(row=i, column=23).value)
            opis_company2 = (sheet.cell(row=i, column=24).value)
            company3 = (sheet.cell(row=i, column=25).value)
            opis_company3 = (sheet.cell(row=i, column=26).value)
            company4 = (sheet.cell(row=i, column=27).value)
            opis_company4 = (sheet.cell(row=i, column=28).value)
            nazvanie = (sheet.cell(row=i, column=29).value)
            pens = (sheet.cell(row=i, column=30).value)
            stud = (sheet.cell(row=i, column=31).value)
            drug_gor = (sheet.cell(row=i, column=32).value)
            invalid = (sheet.cell(row=i, column=33).value)
            migrant = (sheet.cell(row=i, column=34).value)
            oformlenie = (sheet.cell(row=i, column=35).value)
            galka = (sheet.cell(row=i, column=36).value)
            spec1 = (sheet.cell(row=i, column=37).value)
            spec2 = (sheet.cell(row=i, column=38).value)
            #print(vacancy)
            error_string = 1099
            ###############    autorith  ###############

            sheet = wb[lst[2]]
            sheet.title
            login = (sheet.cell(row=1, column=2).value)
            password = (sheet.cell(row=2, column=2).value)
            #login = ('shweper@ya.ru')
            #password = ('Fqtawe98')
            ###############    Open Browser and login  ###############

            error_string = 1110

            #chrome_options.add_argument("--no-startup-window")
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
            browser = webdriver.Chrome(chrome_options=options)
            driver = browser

            browser.get('https://nn.rabota.ru/v3_myVacancy.html?action=create&company_registered=true&employer_registered=true')
            error_string = 1119
            #chrome_options = Options()
            #chrome_options.add_argument("--no-startup-window")
            driver.refresh()
            driver.implicitly_wait(1)
            # Логинимся
            login_bar_xpath = '//*[@id="mail"]'
            login_bar = browser.find_element_by_xpath(login_bar_xpath)
            login_bar.send_keys(login)

            pass_bar_xpath = '//*[@id="password"]'
            pass_bar = browser.find_element_by_xpath(pass_bar_xpath)
            pass_bar.send_keys(password)
            # кликкаем войти
            driver.find_element_by_xpath('//*[@id="authForm"]/input[5]').click()

            error_string = 1143
            #print(opisanie)


            ###################################################OLD##################################################

            vak_bar_xpath = '/html/body/div[2]/div[2]/div[2]/div/div[3]/div[2]/div/div/form/div[1]/div[2]/div/table/tbody/tr[1]/td[2]/div[1]/input'

            """try:
                vak_bar_xpath = '/html/body/div[2]/div[2]/div[2]/div/div[3]/div[2]/div/div/form/div[1]/div[2]/div/table/tbody/tr[1]/td[2]/div[1]/input'
                #optimiz = True
            except:
                try:
                    time.sleep(2)
                    vak_bar_xpath = '/html/body/div[2]/div[2]/div[2]/div/div[3]/div[2]/div/div/form/div[1]/div[2]/div/table/tbody/tr[1]/td[2]/div[1]/input'
                except:
                    time.sleep(3)
                    vak_bar_xpath = '/html/body/div[2]/div[2]/div[2]/div/div[3]/div[2]/div/div/form/div[1]/div[2]/div/table/tbody/tr[1]/td[2]/div[1]/input'
            """
            istina = False  # тест
            while istina == False:
                    try:
                        vak_bar = browser.find_element_by_xpath(vak_bar_xpath)
                        istina = True
                    except:
                        time.sleep(1)

            print(vacancy, "vacancy!")
            time.sleep(1)
            vak_bar.send_keys(vacancy)

            try:
                vak_bar.send_keys(Keys.ARROW_DOWN + Keys.ENTER)
            except:
                error_string = 1156
            error_string = 1157
            '''
            driver.find_element_by_xpath('//*[@id="vacancyForm"]/div[1]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/span[1]').click()
            driver.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/label[5]/input[2]').send_keys(vacancy)
            driver.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/label[5]/input[3]').click()
            driver.find_element_by_xpath('//*[@id="jqmContent"]/div/div[3]/button').click()
            '''
            qwer = 1
            qwert = 0
            #while qwer <= 40:
                #print('/html/body/div[4]/div[1]/div/div[2]/div/label['+str(qwer)+']/input[2]')
            try:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[1]/div[2]/div/table/tbody/tr[2]').click()

                while qwert <= 0:

                    try:
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div/label['+str(qwer)+']/input[2]').send_keys(spec1)
                        browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/label['+str(qwer)+']/input[3]').click()
                        qwert +=1
                    except:

                        qwesadqwdsa= 1
                    qwer += 1
                while qwert <= 1:
                    try:
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div/label['+str(qwer)+']/input[2]').send_keys(spec2)
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div/label['+str(qwer)+']/input[3]').click()
                        qwert +=1
                    except:
                            qwesadqwdsa = 2
                    qwer += 1
                try:
                    browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[3]/button').click()
                except:
                    qwesadqwdsa = 3
            except:
                print('Специализации нет')


            if zp_ot != None:

                zp_ot_bar_xpath = '//*[@id="salary_from"]'
                zp_ot_bar = browser.find_element_by_xpath(zp_ot_bar_xpath)
                zp_ot_bar.send_keys(zp_ot)

            if zp_do != None:
                zp_do_bar_xpath = '//*[@id="salary_to"]'
                zp_do_bar = browser.find_element_by_xpath(zp_do_bar_xpath)
                zp_do_bar.send_keys(zp_do)

            #driver.find_element_by_xpath('//*[@id="mce_3"]/button').click() #почему то вылезла ошибка, временно воткнул слип

            error_string = 1167
            ###############     Опыт работы     ###############

            #browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[1]/td[2]').click()
            if opyt == 'менее года':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[4]').click()
            elif opyt == '1 год':
                browser.find_element_by_xpath('//*[ @ id = "offer_experience_year_count"] / option[5]').click()
            elif opyt == '2 года':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[6]').click()
            elif opyt == '3 года':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[7]').click()
            elif opyt == '4 года':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[8]').click()
            elif opyt == '5 лет':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[9]').click()
            elif opyt == '6 лет':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[10]').click()
            elif opyt == '7 лет':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[11]').click()
            elif opyt == '8 лет':
                 browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[12]').click()
            elif opyt == '9 лет':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[13]').click()
            elif opyt == '10 лет':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[14]').click()
            elif opyt == 'не имеет значения':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[2]').click()
            elif opyt == 'без опыта':
                browser.find_element_by_xpath('//*[@id="offer_experience_year_count"]/option[3]').click()

            error_string = 1198
            ###############     Образование     ###############

            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select').click()
            if obraz == "высшее":
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[2]').click()
            elif obraz == "неполное высшее":
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[3]').click()
            elif obraz == "среднее профессиональное":
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[4]').click()
            elif obraz == "среднее":
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[5]').click()
            elif obraz == "любое":
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[2]/div/div/table/tbody/tr[2]/td[2]/div[1]/div/div/div/div/div/select/option[1]').click()

            error_string = 1212
            ###############     Пол     ###############

            #browser.find_element_by_xpath('//*[@id="is_male"]').click()
            if pol == 'Не важно':
                browser.find_element_by_xpath('//*[@id="is_male"]/option[1]').click()
            elif pol == 'Мужской':
                browser.find_element_by_xpath('//*[@id="is_male"]/option[2]').click()
            elif pol == 'Женский':
                browser.find_element_by_xpath('//*[@id="is_male"]/option[3]').click()

            error_string = 1224
            ###############     Доп настройки кандидатов     ###############

            if pens != None:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[3]/div/table/tbody/tr[3]/td[2]/div/div/div[1]/label/input').click()
            if stud != None:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[3]/div/table/tbody/tr[3]/td[2]/div/div/div[2]/label/input').click()
            if drug_gor != None:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[3]/div/table/tbody/tr[3]/td[2]/div/div/div[3]/label/input').click()
            if invalid != None:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[3]/div/table/tbody/tr[3]/td[2]/div/div/div[4]/label/input').click()
            if migrant != None:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[3]/div/table/tbody/tr[3]/td[2]/div/div/div[5]/label/input').click()

            error_string = 1238
            ###############     возраст     ###############

            vozr_ot_bar_xpath = '//*[@id="age_from"]'
            vozr_ot_bar = browser.find_elements_by_xpath(vozr_ot_bar_xpath)[0]
            vozr_ot_bar.send_keys(vozr_ot)

            vozr_do_bar_xpath = '//*[@id="age_to"]'
            vozr_do_bar = browser.find_elements_by_xpath(vozr_do_bar_xpath)[0]
            vozr_do_bar.send_keys(vozr_do)

            error_string = 1249
            ###############    Контактная информация    ###############

            name_hr_bar_xpath = '//*[@id="vacancyForm"]/div[5]/div/div/table/tbody/tr[1]/td[2]/div/input'
            name_hr_bar = browser.find_element_by_xpath(name_hr_bar_xpath)
            name_hr_bar.clear()
            name_hr_bar.send_keys(name_hr)


            mail_bar_xpath = '//*[@id="contact_email"]'
            mail_bar = browser.find_element_by_xpath(mail_bar_xpath)
            mail_bar.clear()
            mail_bar.send_keys(mail)

            error_string = 1263
            ###############    Phone    ###############

            kod_bar_xpath = '//*[@id="phoneContainerCode_"]'
            kod_bar = browser.find_elements_by_xpath(kod_bar_xpath)[0]
            kod_bar.clear()
            kod_bar.send_keys(kod)

            phone_bar_xpath = '//*[@id="phoneContainerNumber_"]'
            phone_bar = browser.find_elements_by_xpath(phone_bar_xpath)[0]
            phone_bar.clear()
            phone_bar.send_keys(phone)
            if koment != None:
                browser.find_element_by_xpath('//*[@id="phoneContainerComment_"]').send_keys(koment)

            error_string =1278
            ###############    время для звонков    ###############

            select = Select(browser.find_element_by_xpath('//*[@id="phoneContainerCallFrom_"]'))
            select.select_by_visible_text(priem_zv_c)

            select = Select(browser.find_element_by_xpath('//*[@id="phoneContainerCallTo_"]'))
            select.select_by_visible_text(priem_zv_do)

            error_string = 1380
            ###############    дни для звонков    ###############
            browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]').click()
            if opyt == 'Рабочие дни':
                browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[1]').click()
            elif opyt == 'Выходные дни':
                browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[2]').click()
            elif opyt == 'Любой день':
                browser.find_element_by_xpath('//*[@id="phoneContainerCallPeriod_"]/option[3]').click()

            error_string =1404
            ###############    адрес работы    ###############

            browser.find_element_by_xpath('//*[@id="addressesList"]/div/a').click()
            # browser.find_element_by_xpath('//*[@id="vacancyAddressPopupLink"]').click()

            browser.find_element_by_xpath('//*[@id="vacancyCityPopupLink"]').click()
            browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[3]/div/p/a').click()
            browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[1]/input').send_keys(script3.goroda_arr[iter_gorod][0])
            time.sleep(1)
            browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[1]/input').send_keys(Keys.ARROW_DOWN + Keys.ENTER)
            # browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[3]/div/input').click()
            browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[3]/div/input').click()

            # browser.find_element_by_xpath('/html/body/div[6]/div[2]/div/div/div[1]/input').sendKeys(Keys.ENTER)

            browser.find_element_by_xpath('//*[@id="vacancyAddressPopupLink"]').click()
            adr_bar_xpath = '//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[1]/td/input[1]'
            adr_bar = browser.find_element_by_xpath(adr_bar_xpath)
            adr_bar.clear()
            #random_adres = random.randint(1, len(script3.goroda_arr[iter_gorod]) - 1)
            random_adres = random.randint(1, script3.kolichestvo_adresov[iter_gorod]-1)
            print('random adres = ', random_adres)
            adr_bar.send_keys(script3.goroda_arr[iter_gorod][0] + ' ' + script3.goroda_arr[iter_gorod][random_adres])
            time.sleep(1)
            adr_bar.send_keys(Keys.ARROW_DOWN)
            adr_bar.send_keys(Keys.ENTER)
            time.sleep(1)
            browser.find_element_by_xpath(
                '//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[1]/td/input[1]').click()
            time.sleep(1)
            browser.find_element_by_xpath(
                '//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[4]/td/input').click()
            time.sleep(1)
            browser.find_element_by_xpath(
                '//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[6]/td/input').click()

            # browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div/div[2]/div/div/table/tbody/tr/td[1]/div/div[2]/div[2]/div[2]/div/a').click()
            # browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div/div[3]/div/input').click()

            # browser.find_element_by_xpath('//*[@id="vacancyAddressPopup"]/div[2]/div[1]/table/tbody/tr[6]/td/input').click()
            # adr_bar.send_keys(Keys.ENTER)
            ###############  anonim company  ###############

            if nazvanie != None:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[2]/tbody/tr[3]/td[2]/div/label/input').click()
                try:
                    browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[2]/tbody/tr[3]/td[2]/span/input').send_keys(nazvanie)
                except:
                    time.sleep(1)
                    browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[2]/tbody/tr[3]/td[2]/span/input').send_keys(nazvanie)

            error_string = 1436
            ###############    Выбор рубрики    ###############

            #browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[1]/tbody/tr/td[2]/div[1]/a').click()
            #print('ВАЖНО!')
            #browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/div[1]/a').click()
            #print('ВАЖНО!11')

            #time.sleep(3)
            #print('ВАЖНО!111')
            browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[1]/tbody/tr/td[2]/div[1]/a').click()
            try:
                browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/div[1]/a').click()
            except:
                time.sleep(1)
            companys = [company, company1, company2, company3, company4]
            opis_companys = [opis_company, opis_company1, opis_company2, opis_company3, opis_company4]
            #print(companys)
            #print(opis_companys)
            iter_companys = 0

            error_string = 1458

            while iter_companys < len(companys):
                if companys[iter_companys] == None:
                    iter_companys = iter_companys + 1

                else:
                    try:
                        time.sleep(1)
                        browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[1]/tbody/tr/td[2]/div[1]/a').click()

                    except:
                        time.sleep(1)
                    rubriks = script3.xpath_rubriks.get(companys[iter_companys], 0)
                    #print(opis_companys[iter_companys])
                   # print(company, "company")
                    #print(rubriks, "rubriks")
                    browser.find_element_by_xpath(rubriks).click()
                    time.sleep(1)

                    xpartt = vibor_podrubrik.get(companys[iter_companys], 0)
                    #print(companys[iter_companys])
                   # print(xpartt, "xpartt")
                    xpartt2 = xpartt.get(opis_companys[iter_companys])

                    print(xpartt2, "xpartt2")
                    browser.find_element_by_xpath(xpartt2).click() # тестовая строка
                    try:
                        browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/div[3]/a[1]').click()
                    except:
                        browser.find_element_by_xpath('//*[@id="jqmContent"]/div/div[2]/div/div/a[1]').click()
                    time.sleep(1)
                    iter_companys = iter_companys + 1
                    #error_companys = companys[iter_companys]
                    #error_opis = opis_company[iter_companys]
                    error_string =1492

            ###############    График работы    ###############

            if grafik == 'полный рабочий день':
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[1]/label/input').click()
            elif grafik =='сменный график':
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[2]/label/input').click()
            elif grafik == 'свободный график':
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[3]/label/input').click()
            elif grafik == 'частичная занятость':
                browser.find_element_by_xpath('/html/body/div[4]/div[2]/div[2]/div/div[3]/div[2]/div/div/form/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[4]/label/input').click()
            elif grafik == 'удаленная работа':
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[5]/label/input').click()
            elif grafik == 'вахта':
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[4]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]/div[6]/label/input').click()

            error_string =1508
            ####################################################

            yslovia = opisanie.partition('Условия:')
            # print(x[1])
            obiaznosti = yslovia[2].partition('Обязанности:')

            lst_obiaznosti = list(obiaznosti)
            lstobi = lst_obiaznosti[0].split('\n', 1)
            del lstobi[0]

            trebovania = obiaznosti[2].partition('Требования:')
            lst_trebovania = list(trebovania)
            lst_treb = lst_trebovania[2].split('\n', 1)
            # print(obiaznosti[0], trebovania[0], trebovania[2])
            print(lst_trebovania)
            treb0 = lst_trebovania[0].split('\n', 1)
            # del treb0[0]
            iframe = browser.find_elements_by_xpath('//*[@id="description_ifr"]')[0]
            driver.switch_to.default_content()
            driver.switch_to.frame(iframe)
            div = browser.find_elements_by_tag_name('p')[0]
            print("YSLOVIE!!!!!!!!!!!!!\n" ,yslovia, "\n" ,yslovia[0] )
            div.send_keys(Keys.ARROW_UP + Keys.ENTER)
            div1 = browser.find_elements_by_tag_name('p')[0]
            div1.send_keys(" " + (Keys.LEFT_CONTROL + "b"))
            div1.send_keys(yslovia[0])
            browser.find_elements_by_tag_name('strong')[0].send_keys(Keys.ENTER)
            ysl = browser.find_elements_by_tag_name('ul')[0]
            print(lst_obiaznosti)
            print(lst_obiaznosti[0])
            print(lstobi)
            ysl.send_keys(treb0)  # То что после ОБЯЗАННОСТИ
            ysl.send_keys(Keys.BACKSPACE)
            trb = browser.find_elements_by_tag_name('ul')[1]

            tr = lst_treb[1].split("\n \n", 1)
            tr1 = tr[0]
            tr2 = tr[1]
            trb.send_keys(tr1)  # То что после ТРЕБОВАНИЯ
            obz = browser.find_elements_by_tag_name('ul')[2]
            obz.send_keys(lstobi)  # То что после УСЛОВИЯ
            obz.send_keys(Keys.BACKSPACE)
            trsplit = tr2.split(' ', 2)
            tr21 = trsplit[0] + ' ' + trsplit[1]
            tr22 = trsplit[2]
            obz.send_keys(Keys.ENTER + Keys.ENTER)
            lastp = browser.find_elements_by_tag_name('p')[-1]
            lastp.send_keys(Keys.LEFT_CONTROL + "b" + tr21)
            #print(tr21)
            #print(tr22)
            #obz.send_keys(tr21)
            lastp.send_keys(' ' + Keys.LEFT_CONTROL + "b")
            lastp.send_keys(tr22)

            #print(tr22)
            #obz.send_keys(tr22)
            #obz.send_keys(Keys.ENTER + Keys.ENTER + tr2)
            lastp = browser.find_elements_by_tag_name('p')[-1]
            # div.send_keys(Keys.SHIFT + Keys.ARROW_DOWN + Keys.ARROW_DOWN + Keys.ARROW_DOWN)
            # div.send_keys(Keys.LEFT_CONTROL + 'b')
            driver.switch_to.default_content()
            # browser.find_element_by_xpath('//*[@id="mce_6"]/button')
            # browser.find_element_by_xpath('//*[@id="mce_6"]/button')
            #time.sleep(40)
            error_string = 1521

            #####################Оформление#####################

            if oformlenie != None:
                browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[2]/tbody/tr[4]/td[2]/div/div/label/input').click()
                time.sleep(1)
                try:
                    if oformlenie == "Шаблон АМБЕР1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[1]/li[1]/a').click()
                    elif oformlenie =="Шаблон Наземное такси1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[1]/li[2]/a').click()
                    elif oformlenie =="Шаблон Наземное такси2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[1]/li[3]/a').click()
                    elif oformlenie =="Шаблон Наземное такси3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[1]/li[4]/a').click()
                    elif oformlenie =="Шаблон АМБЕР2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[1]/li[5]/a').click()
                    elif oformlenie =="Шаблон Рестораны / Питание1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[1]/a').click()
                    elif oformlenie =="Шаблон Транспорт / Логистика1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[2]/a').click()
                    elif oformlenie =="Шаблон Охрана и безопасность1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[3]/a').click()
                    elif oformlenie =="Шаблон Рестораны / Питание2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[4]/a').click()
                    elif oformlenie =="Шаблон Рестораны / Питание3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[5]/a').click()
                    elif oformlenie =="Шаблон Транспорт / Логистика2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[6]/a').click()
                    elif oformlenie =="Шаблон Строительство1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[7]/a').click()
                    elif oformlenie =="Шаблон Производство / Агропром1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[8]/a').click()
                    elif oformlenie =="Шаблон Торговля1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[9]/a').click()
                    elif oformlenie =="Шаблон Туризм1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[10]/a').click()
                    elif oformlenie =="Шаблон Строительство2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[11]/a').click()
                    elif oformlenie =="Шаблон Торговля2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[12]/a').click()
                    elif oformlenie =="Шаблон Строительство3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[13]/a').click()
                    elif oformlenie =="Шаблон Туризм2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[14]/a').click()
                    elif oformlenie =="Шаблон Производство / Агропром2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[15]/a').click()
                    elif oformlenie =="Шаблон Производство / Агропром3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[16]/a').click()
                    elif oformlenie =="Шаблон Производство1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[17]/a').click()
                    elif oformlenie =="Шаблон Цветы1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[18]/a').click()
                    elif oformlenie =="Шаблон Швейное производство1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[19]/a').click()
                    elif oformlenie =="Шаблон Медицина1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[20]/a').click()
                    elif oformlenie =="Шаблон кол-центр1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[21]/a').click()
                    elif oformlenie =="Шаблон Офисный персонал1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[22]/a').click()
                    elif oformlenie =="Шаблон Рестораны / Питание4":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[23]/a').click()
                    elif oformlenie =="Шаблон Торговля3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[24]/a').click()
                    elif oformlenie =="Шаблон Мебельное производство1":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[25]/a').click()
                    elif oformlenie =="Шаблон Работа на АЗС":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[26]/a').click()
                    elif oformlenie =="Шаблон Производство2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[27]/a').click()
                    elif oformlenie =="Шаблон Производство3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[28]/a').click()
                    elif oformlenie =="Шаблон Столярное дело":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[29]/a').click()
                    elif oformlenie =="Шаблон Торговля4":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[30]/a').click()
                    elif oformlenie =="Шаблон Транспорт / Логистика3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[31]/a').click()
                    elif oformlenie =="Шаблон Транспорт / Логистика4":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[32]/a').click()
                    elif oformlenie =="Шаблон Офисный персонал2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[33]/a').click()
                    elif oformlenie =="Шаблон Торговля5":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[34]/a').click()
                    elif oformlenie =="Шаблон Офисный персонал3":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[35]/a').click()
                    elif oformlenie =="Шаблон Офисный персонал4":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[36]/a').click()
                    elif oformlenie =="Шаблон Охрана и безопасность2":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[37]/a').click()
                    elif oformlenie =="Шаблон Офисный персонал5":
                        browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/ul[2]/li[38]/a').click()
                #    browser.find_element_by_xpath('/html/body/div[4]/div[1]/div/div[2]/div[1]/a').click()
                except:

                    #time.sleep(1)
                    if oformlenie == "Шаблон АМБЕР1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[1]/li[1]/a').click()
                    elif oformlenie == "Шаблон Наземное такси1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[1]/li[2]/a').click()
                    elif oformlenie == "Шаблон Наземное такси2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[1]/li[3]/a').click()
                    elif oformlenie == "Шаблон Наземное такси3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[1]/li[4]/a').click()
                    elif oformlenie == "Шаблон АМБЕР2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[1]/li[5]/a').click()
                    elif oformlenie == "Шаблон Рестораны / Питание1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[1]/a').click()
                    elif oformlenie == "Шаблон Транспорт / Логистика1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[2]/a').click()
                    elif oformlenie == "Шаблон Охрана и безопасность1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[3]/a').click()
                    elif oformlenie == "Шаблон Рестораны / Питание2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[4]/a').click()
                    elif oformlenie == "Шаблон Рестораны / Питание3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[5]/a').click()
                    elif oformlenie == "Шаблон Транспорт / Логистика2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[6]/a').click()
                    elif oformlenie == "Шаблон Строительство1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[7]/a').click()
                    elif oformlenie == "Шаблон Производство / Агропром1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[8]/a').click()
                    elif oformlenie == "Шаблон Торговля1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[9]/a').click()
                    elif oformlenie == "Шаблон Туризм1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[10]/a').click()
                    elif oformlenie == "Шаблон Строительство2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[11]/a').click()
                    elif oformlenie == "Шаблон Торговля2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[12]/a').click()
                    elif oformlenie == "Шаблон Строительство3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[13]/a').click()
                    elif oformlenie == "Шаблон Туризм2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[14]/a').click()
                    elif oformlenie == "Шаблон Производство / Агропром2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[15]/a').click()
                    elif oformlenie == "Шаблон Производство / Агропром3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[16]/a').click()
                    elif oformlenie == "Шаблон Производство1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[17]/a').click()
                    elif oformlenie == "Шаблон Цветы1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[18]/a').click()
                    elif oformlenie == "Шаблон Швейное производство1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[19]/a').click()
                    elif oformlenie == "Шаблон Медицина1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[20]/a').click()
                    elif oformlenie == "Шаблон кол-центр1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[21]/a').click()
                    elif oformlenie == "Шаблон Офисный персонал1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[22]/a').click()
                    elif oformlenie == "Шаблон Рестораны / Питание4":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[23]/a').click()
                    elif oformlenie == "Шаблон Торговля3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[24]/a').click()
                    elif oformlenie == "Шаблон Мебельное производство1":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[25]/a').click()
                    elif oformlenie == "Шаблон Работа на АЗС":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[26]/a').click()
                    elif oformlenie == "Шаблон Производство2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[27]/a').click()
                    elif oformlenie == "Шаблон Производство3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[28]/a').click()
                    elif oformlenie == "Шаблон Столярное дело":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[29]/a').click()
                    elif oformlenie == "Шаблон Торговля4":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[30]/a').click()
                    elif oformlenie == "Шаблон Транспорт / Логистика3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[31]/a').click()
                    elif oformlenie == "Шаблон Транспорт / Логистика4":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[32]/a').click()
                    elif oformlenie == "Шаблон Офисный персонал2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[33]/a').click()
                    elif oformlenie == "Шаблон Торговля5":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[34]/a').click()
                    elif oformlenie == "Шаблон Офисный персонал3":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[35]/a').click()
                    elif oformlenie == "Шаблон Офисный персонал4":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[36]/a').click()
                    elif oformlenie == "Шаблон Охрана и безопасность2":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[37]/a').click()
                    elif oformlenie == "Шаблон Офисный персонал5":
                        browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/ul[2]/li[38]/a').click()
                browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[2]/div[1]/a').click()

            ######################Галка на автоподбор#################
            if galka != None:
                browser.find_element_by_xpath('//*[@id="publication_settings"]/div[2]/label/input').click()

            else:
                time.sleep(0)


            ######################Публикация#################
            #browser.find_element_by_xpath('//*[@id="publication_settings"]/div[2]/label/input').click()
            error_string = 1524
            #browser.find_element_by_xpath('//*[@id="vacancyForm"]/div[6]/div/div/table[2]/tbody/tr[4]/td[2]/div/div[1]/label/input').click()

            #browser.find_element_by_xpath('//*[@id="jqmContent"]/div/a').click()
            time.sleep(1)
            browser.find_element_by_xpath('//*[@id="publishButton"]').click()
            #time.sleep(10)

            # выбор тарифа на тесте не работает, только на прод
            browser.find_element_by_xpath('/html/body/div[2]/div[2]/div[2]/div[1]/div[4]/div[1]/form/div/div[2]/div[1]/div/div/div[2]/div[1]/div/table/tbody/tr[2]/td/div/div[2]/div/label[1]/span[1]').click()

            #######################################################################################################################
            #######################РАЗМЕСТИТЬ ВАКАНСИЮ##############################################################################
            try:
                browser.find_element_by_xpath('//*[@id="vacancyPublishForm"]/div/div[2]/div[2]/a').click()
            except:
                browser.find_element_by_xpath('/html/body/div[2]/div[2]/div[2]/div[1]/div[4]/div[1]/form/div/div[2]/div[2]/a').click()
            browser.quit()
            print(iteracia)
            #except:
            #    time.sleep(1)
            '''
                    log_file = open("error_log.txt", "a")
                    try:
                        log_file.write("ERROR " + str(iteracia) + "\n")
                        log_file.write("Сломалось на строке" + str(error_string) + "\n")
                
                        log_file.write("Город " + script3.goroda_arr[iter_gorod][0] + "\n")
                        log_file.write("Адрес  " + script3.goroda_arr[iter_gorod][random_adres] + "\n")
                
                        log_file.write("Рубрика " + error_companys + "\n")
                        log_file.write("xpath рубрики " + rubriks + "\n")
                        log_file.write("Подрубрика(профессия) " + error_opis + "\n")
                        log_file.write("xpath подрубрики " + xpartt2 + "\n")
                        #log_file.write("Рубрика ",,"\n")
                    except:
                        log_file.write("END_LOG")
                        log_file.close()
                    '''
            browser.quit()
            iter_gorod = iter_gorod + 1
            vsego = vsego + 1

        iter_gorod = 0
        iteracia = iteracia + 1
        print(iter_gorod)
