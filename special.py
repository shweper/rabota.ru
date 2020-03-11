import time

from selenium.webdriver import ActionChains

from selenium import webdriver
from openpyxl import load_workbook
import random
import xlrd
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
browser = webdriver.Chrome(chrome_options=options)
driver = browser

browser.get('https://nn.rabota.ru/v3_myVacancy.html?action=create&company_registered=true&employer_registered=true')
error_string = 1119
# chrome_options = Options()
# chrome_options.add_argument("--no-startup-window")
driver.refresh()
driver.implicitly_wait(1)


wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)
sheet = wb[lst[2]]
sheet.title
login = (sheet.cell(row=1, column=2).value)
password = (sheet.cell(row=2, column=2).value)
# Логинимся
login_bar_xpath = '//*[@id="mail"]'
login_bar = browser.find_element_by_xpath(login_bar_xpath)
login_bar.send_keys(login)

pass_bar_xpath = '//*[@id="password"]'
pass_bar = browser.find_element_by_xpath(pass_bar_xpath)
pass_bar.send_keys(password)
# кликкаем войти
driver.find_element_by_xpath('//*[@id="authForm"]/input[5]').click()
browser.get('https://nn.rabota.ru/v3_myVacancy.html?action=create&company_registered=true&employer_registered=true')
driver.find_element_by_xpath('//*[@id="custom_position"]').send_keys("Главный менеджер по закупкам" + Keys.ENTER)
time.sleep(1)
#Для подключения в основной скрипт, копировить со следующей строчки и до конца.
#переходим на специализации
driver.find_element_by_xpath('//*[@id="vacancyForm"]/div[1]/div[2]/div/table/tbody/tr[2]/td[2]/div[1]').click()
#читаем все места где можно поставить галочку
itemss = driver.find_elements_by_class_name('specialization-item')
#ставим галочку
itemss[15].click()